#!/usr/bin/env python3
"""Test suite for file format handling.

Tests routing logic, direct extraction for text-based formats,
and PDF conversion detection for visual formats.
"""

import sys
import tempfile
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from glm_hybrid_ocr.utils.convert import (
    SUPPORTED_EXTENSIONS,
    VISUAL_EXTENSIONS,
    ensure_pdf,
    is_supported,
)
from glm_hybrid_ocr.utils.extract import (
    DIRECT_EXTRACT_EXTENSIONS,
    extract_to_markdown,
    is_direct_extract,
)


# ---------------------------------------------------------------------------
# Routing tests — verify each extension is classified correctly
# ---------------------------------------------------------------------------


class TestRouting:
    """Ensure every supported extension routes to the right subsystem."""

    @pytest.mark.parametrize("ext", sorted(VISUAL_EXTENSIONS))
    def test_visual_formats_not_direct_extract(self, ext):
        assert not is_direct_extract(f"file{ext}")

    @pytest.mark.parametrize("ext", sorted(DIRECT_EXTRACT_EXTENSIONS))
    def test_direct_formats_are_direct_extract(self, ext):
        assert is_direct_extract(f"file{ext}")

    @pytest.mark.parametrize("ext", sorted(SUPPORTED_EXTENSIONS))
    def test_all_supported_formats_accepted(self, ext):
        assert is_supported(f"file{ext}")

    def test_unsupported_format_rejected(self):
        assert not is_supported("file.xyz")
        assert not is_supported("file.mp4")
        assert not is_supported("file.zip")

    def test_case_insensitive(self):
        assert is_supported("FILE.PDF")
        assert is_supported("data.CSV")
        assert is_supported("report.DOCX")
        assert is_direct_extract("data.CSV")
        assert not is_direct_extract("report.DOCX")

    def test_no_overlap_between_visual_and_direct(self):
        overlap = VISUAL_EXTENSIONS & DIRECT_EXTRACT_EXTENSIONS
        assert overlap == set(), f"Overlapping extensions: {overlap}"

    def test_supported_equals_union(self):
        assert SUPPORTED_EXTENSIONS == VISUAL_EXTENSIONS | DIRECT_EXTRACT_EXTENSIONS


# ---------------------------------------------------------------------------
# Direct extraction tests — TXT
# ---------------------------------------------------------------------------


class TestTxtExtraction:
    def test_plain_text(self, tmp_path):
        f = tmp_path / "test.txt"
        f.write_text("Hello, world!\nSecond line.", encoding="utf-8")
        result = extract_to_markdown(f)
        assert "Hello, world!" in result
        assert "Second line." in result

    def test_empty_file(self, tmp_path):
        f = tmp_path / "empty.txt"
        f.write_text("", encoding="utf-8")
        result = extract_to_markdown(f)
        assert result == ""

    def test_utf8_bom(self, tmp_path):
        f = tmp_path / "bom.txt"
        f.write_bytes(b"\xef\xbb\xbfBOM content here")
        result = extract_to_markdown(f)
        assert "BOM content here" in result

    def test_japanese_text(self, tmp_path):
        f = tmp_path / "jp.txt"
        f.write_text("日本語テスト\n二行目", encoding="utf-8")
        result = extract_to_markdown(f)
        assert "日本語テスト" in result
        assert "二行目" in result

    def test_shift_jis_encoding(self, tmp_path):
        f = tmp_path / "sjis.txt"
        f.write_bytes("日本語テスト".encode("shift_jis"))
        result = extract_to_markdown(f)
        assert "日本語テスト" in result

    def test_multiline_preserves_structure(self, tmp_path):
        content = "Line 1\nLine 2\n\nLine 4 after blank"
        f = tmp_path / "multi.txt"
        f.write_text(content, encoding="utf-8")
        result = extract_to_markdown(f)
        assert result == content


# ---------------------------------------------------------------------------
# Direct extraction tests — CSV
# ---------------------------------------------------------------------------


class TestCsvExtraction:
    def test_basic_csv(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("Name,Age,City\nAlice,30,Tokyo\nBob,25,Osaka", encoding="utf-8")
        result = extract_to_markdown(f)
        assert "| Name | Age | City |" in result
        assert "| --- | --- | --- |" in result
        assert "| Alice | 30 | Tokyo |" in result
        assert "| Bob | 25 | Osaka |" in result

    def test_single_column(self, tmp_path):
        f = tmp_path / "single.csv"
        f.write_text("Items\nApple\nBanana", encoding="utf-8")
        result = extract_to_markdown(f)
        assert "| Items |" in result
        assert "| Apple |" in result

    def test_empty_csv(self, tmp_path):
        f = tmp_path / "empty.csv"
        f.write_text("", encoding="utf-8")
        result = extract_to_markdown(f)
        assert result == ""

    def test_csv_with_commas_in_values(self, tmp_path):
        f = tmp_path / "quoted.csv"
        f.write_text('Name,Description\nAlice,"Hello, world"\nBob,"Foo, bar"', encoding="utf-8")
        result = extract_to_markdown(f)
        assert "Hello, world" in result

    def test_header_only(self, tmp_path):
        f = tmp_path / "header.csv"
        f.write_text("A,B,C\n", encoding="utf-8")
        result = extract_to_markdown(f)
        assert "| A | B | C |" in result
        assert "| --- | --- | --- |" in result

    def test_uneven_rows_padded(self, tmp_path):
        f = tmp_path / "uneven.csv"
        f.write_text("A,B,C\n1\n2,3", encoding="utf-8")
        result = extract_to_markdown(f)
        # Short rows should be padded with empty cells
        lines = result.strip().split("\n")
        assert len(lines) == 4  # header + separator + 2 data rows
        # Each data row should have 3 columns
        for line in lines[2:]:
            assert line.count("|") == 4  # | a | b | c |


# ---------------------------------------------------------------------------
# Direct extraction tests — XLSX
# ---------------------------------------------------------------------------


class TestXlsxExtraction:
    def test_basic_xlsx(self, tmp_path):
        import openpyxl

        f = tmp_path / "data.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append(["Product", "Qty", "Price"])
        ws.append(["Widget", 10, 9.99])
        ws.append(["Gadget", 5, 19.99])
        wb.save(f)

        result = extract_to_markdown(f)
        assert "## Sales" in result
        assert "| Product | Qty | Price |" in result
        assert "Widget" in result
        assert "Gadget" in result

    def test_multiple_sheets(self, tmp_path):
        import openpyxl

        f = tmp_path / "multi.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A", "B"])
        ws1.append([1, 2])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["X", "Y"])
        ws2.append([3, 4])
        wb.save(f)

        result = extract_to_markdown(f)
        assert "## Sheet1" in result
        assert "## Sheet2" in result
        assert "| A | B |" in result
        assert "| X | Y |" in result

    def test_empty_sheet_skipped(self, tmp_path):
        import openpyxl

        f = tmp_path / "empty_sheet.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1.append(["Col1"])
        ws1.append(["Val1"])

        wb.create_sheet("Empty")
        wb.save(f)

        result = extract_to_markdown(f)
        assert "## Data" in result
        assert "## Empty" not in result

    def test_cells_with_none(self, tmp_path):
        import openpyxl

        f = tmp_path / "sparse.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sparse"
        ws.append(["A", "B", "C"])
        ws.append([1, None, 3])
        ws.append([None, 2, None])
        wb.save(f)

        result = extract_to_markdown(f)
        assert "## Sparse" in result
        # None cells should become empty strings
        assert "| 1 |  | 3 |" in result
        assert "|  | 2 |  |" in result


# ---------------------------------------------------------------------------
# Direct extraction tests — HTML
# ---------------------------------------------------------------------------


class TestHtmlExtraction:
    def test_basic_html(self, tmp_path):
        f = tmp_path / "page.html"
        f.write_text(
            "<html><body><h1>Title</h1><p>Hello world</p></body></html>",
            encoding="utf-8",
        )
        result = extract_to_markdown(f)
        assert "# Title" in result
        assert "Hello world" in result

    def test_nested_headings(self, tmp_path):
        f = tmp_path / "headings.html"
        f.write_text(
            "<h1>H1</h1><h2>H2</h2><h3>H3</h3>",
            encoding="utf-8",
        )
        result = extract_to_markdown(f)
        assert "# H1" in result
        assert "## H2" in result
        assert "### H3" in result

    def test_script_and_style_stripped(self, tmp_path):
        f = tmp_path / "scripts.html"
        f.write_text(
            "<html><head><style>body{color:red}</style></head>"
            "<body><script>alert('hi')</script><p>Visible text</p></body></html>",
            encoding="utf-8",
        )
        result = extract_to_markdown(f)
        assert "Visible text" in result
        assert "alert" not in result
        assert "color:red" not in result

    def test_htm_extension(self, tmp_path):
        f = tmp_path / "page.htm"
        f.write_text("<p>HTM content</p>", encoding="utf-8")
        result = extract_to_markdown(f)
        assert "HTM content" in result

    def test_blank_lines_collapsed(self, tmp_path):
        f = tmp_path / "spaced.html"
        f.write_text(
            "<p>Para 1</p><p></p><p></p><p>Para 2</p>",
            encoding="utf-8",
        )
        result = extract_to_markdown(f)
        assert "Para 1" in result
        assert "Para 2" in result
        # Should not have more than one consecutive blank line
        assert "\n\n\n" not in result

    def test_list_items(self, tmp_path):
        f = tmp_path / "list.html"
        f.write_text(
            "<ul><li>Item 1</li><li>Item 2</li><li>Item 3</li></ul>",
            encoding="utf-8",
        )
        result = extract_to_markdown(f)
        assert "Item 1" in result
        assert "Item 2" in result
        assert "Item 3" in result


# ---------------------------------------------------------------------------
# PDF conversion routing tests
# ---------------------------------------------------------------------------


class TestPdfConversion:
    def test_pdf_returns_same_path(self, tmp_path):
        f = tmp_path / "doc.pdf"
        f.write_bytes(b"%PDF-1.4 fake")
        result = ensure_pdf(f)
        assert result == f

    def test_direct_extract_format_raises(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("a,b\n1,2", encoding="utf-8")
        with pytest.raises(ValueError, match="does not need PDF conversion"):
            ensure_pdf(f)

    def test_unsupported_extension_raises(self, tmp_path):
        f = tmp_path / "file.mp4"
        f.write_bytes(b"fake")
        with pytest.raises(ValueError):
            ensure_pdf(f)


# ---------------------------------------------------------------------------
# Unsupported format tests
# ---------------------------------------------------------------------------


class TestUnsupportedFormats:
    def test_extract_unsupported_raises(self, tmp_path):
        f = tmp_path / "file.pdf"
        f.write_bytes(b"%PDF")
        with pytest.raises(ValueError, match="No direct extractor"):
            extract_to_markdown(f)

    def test_extract_unknown_ext_raises(self, tmp_path):
        f = tmp_path / "file.xyz"
        f.write_bytes(b"data")
        with pytest.raises(ValueError):
            extract_to_markdown(f)


# ---------------------------------------------------------------------------
# Integration: verify extract_to_markdown dispatches correctly for each type
# ---------------------------------------------------------------------------


class TestExtractionDispatch:
    """Verify extract_to_markdown handles each direct-extract extension."""

    def test_txt_dispatches(self, tmp_path):
        f = tmp_path / "test.txt"
        f.write_text("hello", encoding="utf-8")
        assert extract_to_markdown(f) == "hello"

    def test_csv_dispatches(self, tmp_path):
        f = tmp_path / "test.csv"
        f.write_text("a,b\n1,2", encoding="utf-8")
        result = extract_to_markdown(f)
        assert "| a | b |" in result

    def test_xlsx_dispatches(self, tmp_path):
        import openpyxl

        f = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        wb.active.append(["x"])
        wb.active.append([1])
        wb.save(f)
        result = extract_to_markdown(f)
        assert "| x |" in result

    def test_html_dispatches(self, tmp_path):
        f = tmp_path / "test.html"
        f.write_text("<p>hi</p>", encoding="utf-8")
        assert "hi" in extract_to_markdown(f)

    def test_htm_dispatches(self, tmp_path):
        f = tmp_path / "test.htm"
        f.write_text("<p>hi</p>", encoding="utf-8")
        assert "hi" in extract_to_markdown(f)
