"""Direct text extraction for text-based file formats.

These formats contain machine-readable content that doesn't need OCR.
Content is extracted directly and converted to markdown.
"""

import csv
import html.parser
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# Formats where direct extraction is better than OCR
DIRECT_EXTRACT_EXTENSIONS = {
    ".txt", ".csv",
    ".xlsx", ".xls", ".ods",
    ".html", ".htm",
}


def is_direct_extract(file_path: str | Path) -> bool:
    """Check if the file should be directly extracted instead of OCR'd."""
    return Path(file_path).suffix.lower() in DIRECT_EXTRACT_EXTENSIONS


def extract_to_markdown(file_path: str | Path) -> str:
    """Extract content from a text-based file and return as markdown.

    Args:
        file_path: Path to the input file.

    Returns:
        Markdown string with the extracted content.

    Raises:
        ValueError: If the file type is not supported for direct extraction.
    """
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()

    if suffix == ".txt":
        return _extract_txt(file_path)
    elif suffix == ".csv":
        return _extract_csv(file_path)
    elif suffix in (".xlsx", ".xls", ".ods"):
        return _extract_spreadsheet(file_path)
    elif suffix in (".html", ".htm"):
        return _extract_html(file_path)
    else:
        raise ValueError(f"No direct extractor for: {suffix}")


def _extract_txt(file_path: Path) -> str:
    """Read plain text file."""
    for encoding in ("utf-8", "utf-8-sig", "shift_jis", "cp932", "euc-jp", "latin-1"):
        try:
            return file_path.read_text(encoding=encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return file_path.read_text(encoding="utf-8", errors="replace")


def _extract_csv(file_path: Path) -> str:
    """Convert CSV to markdown table."""
    text = _extract_txt(file_path)
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    if not rows:
        return ""

    header = rows[0]
    md = "| " + " | ".join(header) + " |\n"
    md += "| " + " | ".join("---" for _ in header) + " |\n"
    for row in rows[1:]:
        # Pad or truncate to match header length
        padded = row + [""] * (len(header) - len(row))
        md += "| " + " | ".join(padded[:len(header)]) + " |\n"

    return md


def _extract_spreadsheet(file_path: Path) -> str:
    """Convert spreadsheet (xlsx/xls/ods) to markdown tables."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for spreadsheet extraction. "
            "Install it with: pip install openpyxl"
        )

    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    elif suffix in (".xls", ".ods"):
        # For xls/ods, try converting via csv as fallback
        try:
            import pandas as pd
            sheets_md = []
            xls = pd.ExcelFile(file_path)
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                if df.empty:
                    continue
                sheets_md.append(f"## {sheet_name}\n\n{df.to_markdown(index=False)}")
            return "\n\n".join(sheets_md) if sheets_md else ""
        except ImportError:
            raise RuntimeError(
                f"pandas is required for {suffix} extraction. "
                "Install it with: pip install pandas openpyxl"
            )
    else:
        raise ValueError(f"Unsupported spreadsheet format: {suffix}")

    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find the actual data range (skip fully empty rows)
        data_rows = [r for r in rows if any(c is not None for c in r)]
        if not data_rows:
            continue

        num_cols = max(len(r) for r in data_rows)
        header = data_rows[0]
        header_strs = [str(c) if c is not None else "" for c in header]
        # Pad to num_cols
        header_strs += [""] * (num_cols - len(header_strs))

        md = f"## {sheet_name}\n\n"
        md += "| " + " | ".join(header_strs) + " |\n"
        md += "| " + " | ".join("---" for _ in range(num_cols)) + " |\n"

        for row in data_rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            cells += [""] * (num_cols - len(cells))
            md += "| " + " | ".join(cells[:num_cols]) + " |\n"

        sections.append(md)

    wb.close()
    return "\n\n".join(sections) if sections else ""


class _HTMLTextExtractor(html.parser.HTMLParser):
    """Simple HTML to text converter."""

    def __init__(self):
        super().__init__()
        self._parts: list[str] = []
        self._skip = False

    def handle_starttag(self, tag, attrs):
        if tag in ("script", "style"):
            self._skip = True
        elif tag in ("p", "div", "br", "h1", "h2", "h3", "h4", "h5", "h6", "li", "tr"):
            self._parts.append("\n")
        if tag in ("h1", "h2", "h3", "h4", "h5", "h6"):
            level = int(tag[1])
            self._parts.append("#" * level + " ")

    def handle_endtag(self, tag):
        if tag in ("script", "style"):
            self._skip = False
        elif tag in ("p", "div", "h1", "h2", "h3", "h4", "h5", "h6", "tr"):
            self._parts.append("\n")

    def handle_data(self, data):
        if not self._skip:
            self._parts.append(data)

    def get_text(self) -> str:
        text = "".join(self._parts)
        # Collapse multiple blank lines
        lines = text.split("\n")
        result = []
        prev_blank = False
        for line in lines:
            stripped = line.strip()
            if not stripped:
                if not prev_blank:
                    result.append("")
                prev_blank = True
            else:
                result.append(stripped)
                prev_blank = False
        return "\n".join(result).strip()


def _extract_html(file_path: Path) -> str:
    """Extract text content from HTML."""
    text = _extract_txt(file_path)
    parser = _HTMLTextExtractor()
    parser.feed(text)
    return parser.get_text()
